"""Reproduce the Zarattini/Barbon/Aziz 5-minute ORB "Stocks in Play" study on
NSE (NIFTY 200) and write a paper-comparable Excel report.

    python -m app.orb_report --start 2023-01-01 --end 2024-12-31 \
        --timeframe 5m --capital 1000000 --slippage-bps 10 --out orb_nifty200.xlsx

What it does that the standard backtest report does not:
  * per-trade R multiple  (R = atr_stop_fraction x daily ATR14 at entry)
  * per-trade opening-range Relative Volume, and the Figure-4 bucket table
    (avg R vs RVOL) — the single best "does the edge port to NSE" diagnostic
  * daily-frequency Sharpe / vol / MDD and an alpha/beta regression of the
    strategy's daily returns on NIFTY 50 (paper Table 2 format)

Data comes from the connected Zerodha session (chunked historical fetch, so
the multi-year 5-minute range is stitched from successive Kite pages) and is
cached on disk, so re-runs are fast.
"""

from __future__ import annotations

import argparse
import math
import statistics
from collections import defaultdict, deque
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.backtesting.costs import CostConfig, CostModel
from app.backtesting.engine import BacktestEngine
from app.backtesting.timeframes import kite_interval, resolve
from app.backtesting.trades import ClosedTrade, reconstruct_trades
from app.config import get_settings
from app.db.session import SessionLocal
from app.market_data.instruments import resolve_instrument_token
from app.market_data.nse_universe import NIFTY_200
from app.services import broker_service
from app.strategies.base import Bar
from app.strategies.library import OpeningBreakoutUSStrategy

_ATR_PERIOD = 14
_RVOL_LOOKBACK = 14
_ATR_STOP_FRACTION = 0.10  # matches the "balanced" preset / the paper


# --------------------------------------------------------------------------
# data
# --------------------------------------------------------------------------

def _to_dt(ts: Any) -> datetime:
    if isinstance(ts, datetime):
        return ts
    s = str(ts).strip().replace("Z", "+00:00")
    if len(s) >= 5 and s[-5] in "+-" and s[-3] != ":":
        s = s[:-2] + ":" + s[-2:]
    return datetime.fromisoformat(s)


def _fetch(client: Any, symbols: list[str], interval: str, frm: datetime, to: datetime) -> tuple[
    dict[str, list[Bar]], list[str]
]:
    out: dict[str, list[Bar]] = {}
    skipped: list[str] = []
    for i, sym in enumerate(symbols, 1):
        ref = f"NSE:{sym}"
        try:
            token, ts = resolve_instrument_token(ref)
        except Exception:
            skipped.append(sym)
            continue
        try:
            rows = client.get_historical_candles(token, interval, frm, to)
        except Exception as exc:  # noqa: BLE001
            print(f"  ! {sym}: fetch failed ({exc})")
            skipped.append(sym)
            continue
        if not rows:
            skipped.append(sym)
            continue
        out[ts] = [
            Bar(timestamp=r[0], open=r[1], high=r[2], low=r[3], close=r[4],
                volume=r[5] if len(r) > 5 else 0, instrument=ts)
            for r in rows
        ]
        if i % 25 == 0:
            print(f"  fetched {i}/{len(symbols)} …")
    return out, skipped


def _daily_bars(client: Any, symbol: str, frm: datetime, to: datetime) -> list[list[Any]]:
    token, _ts = resolve_instrument_token(f"NSE:{symbol}")
    return client.get_historical_candles(token, "day", frm - timedelta(days=40), to)


# --------------------------------------------------------------------------
# per-instrument analytics for R and RVOL
# --------------------------------------------------------------------------

def _wilder_atr_by_date(daily_rows: list[list[Any]], period: int) -> dict[date, float]:
    """ATR14 available *before* each session's open (uses bars strictly prior)."""
    if len(daily_rows) < period + 2:
        return {}
    highs = [float(r[2]) for r in daily_rows]
    lows = [float(r[3]) for r in daily_rows]
    closes = [float(r[4]) for r in daily_rows]
    dates = [_to_dt(r[0]).date() for r in daily_rows]
    trs = [
        max(highs[i] - lows[i], abs(highs[i] - closes[i - 1]), abs(lows[i] - closes[i - 1]))
        for i in range(1, len(daily_rows))
    ]
    out: dict[date, float] = {}
    atr = sum(trs[:period]) / period
    # atr computed through index `period` (bars 1..period) is known before bar period+1's open
    for i in range(period, len(trs)):
        out[dates[i + 1]] = atr
        atr = (atr * (period - 1) + trs[i]) / period
    return out


def _or_rvol_by_date(bars_5m: list[Bar], lookback: int) -> dict[date, float]:
    """Opening-range (09:15 bar) volume / trailing mean of it."""
    or_vol: dict[date, float] = {}
    for b in bars_5m:
        dt = _to_dt(b.timestamp)
        if dt.hour == 9 and dt.minute == 15:
            or_vol[dt.date()] = float(b.volume or 0.0)
    out: dict[date, float] = {}
    hist: deque[float] = deque(maxlen=lookback)
    for d in sorted(or_vol):
        if len(hist) == lookback:
            avg = sum(hist) / lookback
            if avg > 0:
                out[d] = or_vol[d] / avg
        hist.append(or_vol[d])
    return out


# --------------------------------------------------------------------------
# metrics
# --------------------------------------------------------------------------

def _daily_nav(equity_curve: list[tuple[Any, float]]) -> list[tuple[date, float]]:
    by_day: dict[date, float] = {}
    for ts, v in equity_curve:
        by_day[_to_dt(ts).date()] = float(v)
    return sorted(by_day.items())


def _returns(nav: list[float]) -> list[float]:
    # only while solvent; clamp each day to -100% so a zero-crossing NAV
    # can't emit a -441% "daily return" that wrecks vol / Sharpe.
    return [
        max(-1.0, nav[i] / nav[i - 1] - 1.0)
        for i in range(1, len(nav))
        if nav[i - 1] > 0 and nav[i] > 0
    ]


def _mdd(nav: list[float]) -> float:
    peak = -math.inf
    mdd = 0.0
    for v in nav:
        peak = max(peak, v)
        if peak > 0:
            mdd = max(mdd, (peak - v) / peak)
    return mdd


def _ols(y: list[float], x: list[float]) -> tuple[float, float, float]:
    """Return (alpha_per_period, beta, r2) for y = alpha + beta*x."""
    n = len(y)
    if n < 10:
        return 0.0, 0.0, 0.0
    mx = sum(x) / n
    my = sum(y) / n
    sxx = sum((xi - mx) ** 2 for xi in x)
    sxy = sum((x[i] - mx) * (y[i] - my) for i in range(n))
    if sxx == 0:
        return 0.0, 0.0, 0.0
    beta = sxy / sxx
    alpha = my - beta * mx
    ss_tot = sum((yi - my) ** 2 for yi in y)
    ss_res = sum((y[i] - (alpha + beta * x[i])) ** 2 for i in range(n))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0
    return alpha, beta, r2


def _perf_block(nav: list[tuple[date, float]], bench_ret_by_date: dict[date, float]) -> dict[str, Any]:
    vals = [v for _d, v in nav]
    rets = _returns(vals)
    dsr = [d for d, _v in nav][1:]
    ann = 252
    total = max(-1.0, vals[-1] / vals[0] - 1.0) if vals and vals[0] > 0 else 0.0
    years = max(1e-9, (nav[-1][0] - nav[0][0]).days / 365.25)
    cagr = (vals[-1] / vals[0]) ** (1 / years) - 1.0 if vals and vals[-1] > 0 and vals[0] > 0 else -1.0
    vol = statistics.pstdev(rets) * math.sqrt(ann) if len(rets) > 1 else 0.0
    mean = statistics.fmean(rets) if rets else 0.0
    sharpe = (mean * ann) / vol if vol > 0 else 0.0
    downs = [r for r in rets if r < 0]
    dd = statistics.pstdev(downs) * math.sqrt(ann) if len(downs) > 1 else 0.0
    sortino = (mean * ann) / dd if dd > 0 else 0.0
    mdd = _mdd(vals)
    # align with benchmark
    raw_pairs = [(rets[i], bench_ret_by_date.get(dsr[i])) for i in range(len(rets))]
    pairs: list[tuple[float, float]] = [(a, b) for a, b in raw_pairs if b is not None]
    alpha_d, beta, r2 = _ols([a for a, _b in pairs], [b for _a, b in pairs]) if pairs else (0, 0, 0)
    return {
        "total_return_pct": total * 100,
        "cagr_pct": cagr * 100,
        "ann_vol_pct": vol * 100,
        "sharpe": sharpe,
        "sortino": sortino,
        "mdd_pct": mdd * 100,
        "worst_day_pct": (min(rets) * 100) if rets else 0.0,
        "best_day_pct": (max(rets) * 100) if rets else 0.0,
        "alpha_ann_pct": alpha_d * ann * 100,
        "beta": beta,
        "r2": r2,
        "days": len(vals),
    }


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

def run(args: argparse.Namespace) -> None:
    tf = resolve(args.timeframe)
    frm = datetime.fromisoformat(args.start)
    to = datetime.fromisoformat(args.end)
    capital = float(args.capital)

    db = SessionLocal()
    try:
        client = broker_service.build_authenticated_client(db, get_settings())
    finally:
        db.close()

    universe = NIFTY_200[: args.limit] if args.limit else NIFTY_200
    print(f"Universe: NIFTY 200 ({len(universe)} names) | {args.start} → {args.end} | {tf.token}")
    print("Fetching intraday candles (chunked; first run is slow, then cached)…")
    candles, skipped = _fetch(client, universe, kite_interval(tf.token), frm, to)
    print(f"  usable instruments: {len(candles)}   skipped: {len(skipped)} {skipped[:10]}"
          f"{' …' if len(skipped) > 10 else ''}")

    print("Fetching daily bars for ATR / RVOL context…")
    atr_by_sym: dict[str, dict[date, float]] = {}
    rvol_by_sym: dict[str, dict[date, float]] = {}
    for sym in list(candles):
        try:
            atr_by_sym[sym] = _wilder_atr_by_date(_daily_bars(client, sym, frm, to), _ATR_PERIOD)
        except Exception:  # noqa: BLE001
            atr_by_sym[sym] = {}
        rvol_by_sym[sym] = _or_rvol_by_date(candles[sym], _RVOL_LOOKBACK)

    print("Fetching NIFTY 50 for the benchmark regression…")
    nifty_daily = client.get_historical_candles(
        resolve_instrument_token("NSE:NIFTY 50")[0], "day", frm, to
    )
    bench_close = {_to_dt(r[0]).date(): float(r[4]) for r in nifty_daily}
    bench_dates = sorted(bench_close)
    bench_ret = {
        bench_dates[i]: bench_close[bench_dates[i]] / bench_close[bench_dates[i - 1]] - 1.0
        for i in range(1, len(bench_dates))
    }
    bench_total = (
        bench_close[bench_dates[-1]] / bench_close[bench_dates[0]] - 1.0 if len(bench_dates) > 1 else 0.0
    )

    print("Running the engine…")
    params = OpeningBreakoutUSStrategy.resolve_params(
        OpeningBreakoutUSStrategy.presets()["balanced"]
    )
    cost_model = CostModel(CostConfig(slippage_bps=float(args.slippage_bps)))
    result = BacktestEngine(
        OpeningBreakoutUSStrategy, params, capital, cost_model=cost_model,
        max_gross_exposure=4.0,  # the paper's leverage ceiling
    ).run(candles)
    if result.diagnostics.ruined:
        print(f"  ⚠ RUINED at {result.diagnostics.ruin_ts}")
    print(f"  peak gross exposure: {result.diagnostics.peak_gross_exposure_pct:.0f}% of capital"
          f"  ({result.diagnostics.exposure_capped_orders} orders scaled by the 4x cap)")

    mark = {s: b[-1].close for s, b in candles.items() if b}
    trades = [
        t for t in reconstruct_trades(
            result.fills, fill_costs=[f.cost for f in result.fills], mark_prices=mark
        )
        if not t.is_open
    ]
    print(f"  fills={result.diagnostics.fills}  closed trades={len(trades)}  "
          f"costs=₹{result.total_costs:,.0f}")

    _write_excel(
        args.out, args, tf.token, capital, len(candles), skipped, trades, result,
        atr_by_sym, rvol_by_sym, bench_ret, bench_total,
    )
    print(f"\nWrote {args.out}")


def _trade_rows(
    trades: list[ClosedTrade],
    atr_by_sym: dict[str, dict[date, float]],
    rvol_by_sym: dict[str, dict[date, float]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for t in trades:
        edt = _to_dt(t.entry_time)
        atr = atr_by_sym.get(t.instrument, {}).get(edt.date())
        r_per_share = _ATR_STOP_FRACTION * atr if atr else None
        signed_move = (t.exit_price - t.entry_price) if t.direction == "long" else (
            t.entry_price - t.exit_price
        )
        r_mult = (signed_move / r_per_share) if r_per_share else None
        net_r = (t.net_pnl / (r_per_share * t.quantity)) if r_per_share and t.quantity else None
        rows.append({
            "instrument": t.instrument,
            "direction": t.direction,
            "entry_time": str(t.entry_time),
            "exit_time": str(t.exit_time),
            "qty": t.quantity,
            "entry": round(t.entry_price, 2),
            "exit": round(t.exit_price, 2),
            "atr14": round(atr, 2) if atr else None,
            "risk_per_share": round(r_per_share, 3) if r_per_share else None,
            "gross_pnl": round(t.gross_pnl, 2),
            "costs": round(t.costs, 2),
            "net_pnl": round(t.net_pnl, 2),
            "return_pct": round(t.return_pct, 3),
            "r_multiple": round(r_mult, 3) if r_mult is not None else None,
            "net_r": round(net_r, 3) if net_r is not None else None,
            "rvol": round(rvol_by_sym.get(t.instrument, {}).get(edt.date(), float("nan")), 2),
            "bars_held": t.bars_held,
        })
    return rows


def _write_excel(
    path: str, args: argparse.Namespace, tf: str, capital: float, n_instruments: int,
    skipped: list[str], trades: list[ClosedTrade], result: Any,
    atr_by_sym: dict[str, dict[date, float]], rvol_by_sym: dict[str, dict[date, float]],
    bench_ret: dict[date, float], bench_total: float,
) -> None:
    wb = Workbook()
    bold = Font(bold=True)

    rows = _trade_rows(trades, atr_by_sym, rvol_by_sym)
    nav = _daily_nav(result.equity_curve)
    perf = _perf_block(nav, bench_ret)

    net_rs = [r["net_r"] for r in rows if r["net_r"] is not None]
    hit = sum(1 for r in rows if r["net_pnl"] > 0) / len(rows) if rows else 0.0

    # ---- Summary ----
    s = wb.active
    s.title = "Summary"
    def put(r: int, label: str, val: Any) -> None:
        s.cell(r, 1, label).font = bold
        s.cell(r, 2, val)
    put(1, "5-minute ORB on NIFTY 200 — Stocks in Play (RVOL) reproduction", "")
    put(3, "Period", f"{args.start} → {args.end}")
    put(4, "Timeframe", tf)
    put(5, "Universe (usable / requested)", f"{n_instruments} / {len(NIFTY_200)}")
    put(6, "Skipped tickers", ", ".join(skipped))
    put(7, "Preset", "balanced (OR 5m, RVOL≥1.0, Top-20, stop 10%×ATR14, EoD exit, 1% risk, MIS)")
    put(8, "Starting capital (₹)", capital)
    put(9, "Slippage (bps/side)", args.slippage_bps)
    put(10, "Commission model", "Zerodha intraday: brokerage + STT + exch + GST + stamp + SEBI")
    put(12, "— Strategy (daily frequency) —", "")
    put(13, "Total return %", round(perf["total_return_pct"], 2))
    put(14, "CAGR %", round(perf["cagr_pct"], 2))
    put(15, "Annualised vol %", round(perf["ann_vol_pct"], 2))
    put(16, "Sharpe (daily, ann.)", round(perf["sharpe"], 2))
    put(17, "Sortino", round(perf["sortino"], 2))
    put(18, "Max drawdown %", round(perf["mdd_pct"], 2))
    put(19, "Worst day %", round(perf["worst_day_pct"], 2))
    put(20, "Best day %", round(perf["best_day_pct"], 2))
    put(21, "Alpha vs NIFTY 50 (%/yr)", round(perf["alpha_ann_pct"], 2))
    put(22, "Beta vs NIFTY 50", round(perf["beta"], 3))
    put(23, "Regression R²", round(perf["r2"], 3))
    put(25, "— Trades —", "")
    put(26, "Closed trades", len(rows))
    put(27, "Hit ratio %", round(hit * 100, 1))
    put(28, "Avg trade (₹)", round(sum(r["net_pnl"] for r in rows) / len(rows), 0) if rows else 0)
    put(29, "Avg trade (R)", round(statistics.fmean(net_rs), 3) if net_rs else None)
    put(30, "Median trade (R)", round(statistics.median(net_rs), 3) if net_rs else None)
    put(31, "Total costs (₹)", round(result.total_costs, 0))
    put(32, "Total net P&L (₹)", round(sum(r["net_pnl"] for r in rows), 0))
    put(34, "— NIFTY 50 buy & hold —", "")
    put(35, "Total return %", round(bench_total * 100, 2))
    s.column_dimensions["A"].width = 42
    s.column_dimensions["B"].width = 60

    # ---- Trades ----
    ts = wb.create_sheet("Trades")
    headers = list(rows[0].keys()) if rows else ["(no trades)"]
    for c, h in enumerate(headers, 1):
        ts.cell(1, c, h).font = bold
    for ri, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ts.cell(ri, c, row[h])
    ts.freeze_panes = "A2"

    # ---- RVOL buckets (Figure 4) ----
    bk = wb.create_sheet("RVOL Buckets")
    edges = [0, 1, 2, 3, 4, 5, 7, 10, 15, 20, 30, 1e9]
    labels = ["<1", "1-2", "2-3", "3-4", "4-5", "5-7", "7-10", "10-15", "15-20", "20-30", ">30"]
    agg: dict[int, list[float]] = defaultdict(list)
    wins: dict[int, int] = defaultdict(int)
    for r in rows:
        rv = r["rvol"]
        nr = r["net_r"]
        if nr is None or rv != rv:  # nan check
            continue
        for i in range(len(edges) - 1):
            if edges[i] <= rv < edges[i + 1]:
                agg[i].append(nr)
                wins[i] += 1 if r["net_pnl"] > 0 else 0
                break
    bk.cell(1, 1, "RVOL bucket").font = bold
    bk.cell(1, 2, "Trades").font = bold
    bk.cell(1, 3, "Avg net P&L (R)").font = bold
    bk.cell(1, 4, "Win %").font = bold
    for i, lab in enumerate(labels):
        vals = agg.get(i, [])
        bk.cell(i + 2, 1, lab)
        bk.cell(i + 2, 2, len(vals))
        bk.cell(i + 2, 3, round(statistics.fmean(vals), 4) if vals else None)
        bk.cell(i + 2, 4, round(wins[i] / len(vals) * 100, 1) if vals else None)
    chart = BarChart()
    chart.title = "Figure 4 — Avg PnL (R) by opening-range Relative Volume"
    chart.y_axis.title = "Avg net PnL (R)"
    chart.x_axis.title = "RVOL bucket"
    data = Reference(bk, min_col=3, min_row=1, max_row=len(labels) + 1)
    cats = Reference(bk, min_col=1, min_row=2, max_row=len(labels) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 18
    bk.add_chart(chart, "F2")

    # ---- Monthly returns ----
    mo = wb.create_sheet("Monthly")
    by_month: dict[tuple[int, int], list[float]] = defaultdict(list)
    for i in range(1, len(nav)):
        d, v = nav[i]
        pv = nav[i - 1][1]
        if pv:
            by_month[(d.year, d.month)].append(v / pv - 1.0)
    years = sorted({y for y, _m in by_month})
    mo.cell(1, 1, "Year").font = bold
    for m in range(1, 13):
        mo.cell(1, m + 1, datetime(2000, m, 1).strftime("%b")).font = bold
    mo.cell(1, 14, "Year %").font = bold
    for ri, y in enumerate(years, 2):
        mo.cell(ri, 1, y)
        comp = 1.0
        for m in range(1, 13):
            rs = by_month.get((y, m))
            if rs:
                mret = math.prod(1 + x for x in rs) - 1.0
                mo.cell(ri, m + 1, round(mret * 100, 2))
                comp *= 1 + mret
        mo.cell(ri, 14, round((comp - 1) * 100, 2))

    # ---- Yearly ----
    yr = wb.create_sheet("Yearly")
    yr.cell(1, 1, "Year").font = bold
    for c, h in enumerate(["Trades", "Hit %", "Net P&L (₹)", "Avg R", "Best R", "Worst R"], 2):
        yr.cell(1, c, h).font = bold
    by_year: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_year[_to_dt(r["entry_time"]).year].append(r)
    for ri, y in enumerate(sorted(by_year), 2):
        rr = by_year[y]
        rs = [x["net_r"] for x in rr if x["net_r"] is not None]
        yr.cell(ri, 1, y)
        yr.cell(ri, 2, len(rr))
        yr.cell(ri, 3, round(sum(1 for x in rr if x["net_pnl"] > 0) / len(rr) * 100, 1))
        yr.cell(ri, 4, round(sum(x["net_pnl"] for x in rr), 0))
        yr.cell(ri, 5, round(statistics.fmean(rs), 3) if rs else None)
        yr.cell(ri, 6, round(max(rs), 2) if rs else None)
        yr.cell(ri, 7, round(min(rs), 2) if rs else None)

    # ---- Equity curve ----
    ec = wb.create_sheet("Equity Curve")
    ec.cell(1, 1, "Date").font = bold
    ec.cell(1, 2, "Strategy NAV").font = bold
    ec.cell(1, 3, "NIFTY 50 (rebased)").font = bold
    base = nav[0][1] if nav else capital
    bench_nav = base
    prev_d = nav[0][0] if nav else None
    for ri, (d, v) in enumerate(nav, 2):
        ec.cell(ri, 1, d.isoformat())
        ec.cell(ri, 2, round(v, 0))
        if prev_d is not None and d != prev_d:
            bench_nav *= 1 + bench_ret.get(d, 0.0)
        prev_d = d
        ec.cell(ri, 3, round(bench_nav, 0))
    lc = LineChart()
    lc.title = "Equity curve — strategy vs NIFTY 50 buy & hold"
    lc.height = 10
    lc.width = 26
    ref = Reference(ec, min_col=2, max_col=3, min_row=1, max_row=len(nav) + 1)
    lc.add_data(ref, titles_from_data=True)
    lc.set_categories(Reference(ec, min_col=1, min_row=2, max_row=len(nav) + 1))
    ec.add_chart(lc, "E2")

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = max(
                12, sheet.column_dimensions[get_column_letter(col)].width or 0
            )

    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--start", default="2023-01-01")
    ap.add_argument("--end", default="2024-12-31")
    ap.add_argument("--timeframe", default="5m")
    ap.add_argument("--capital", type=float, default=1_000_000)
    ap.add_argument("--slippage-bps", type=float, default=10.0)
    ap.add_argument("--out", default="orb_nifty200_report.xlsx")
    ap.add_argument("--limit", type=int, default=0, help="cap the universe (smoke tests)")
    run(ap.parse_args())


if __name__ == "__main__":
    main()
